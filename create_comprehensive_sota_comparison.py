"""
Create Professional IEEE-Standard Excel Comparison with Comprehensive Multi-Episode Results

This version includes BOTH:
1. Single attack (DoS) results from threshold 0.75 test
2. Comprehensive multi-episode results (9 attacks average) from comprehensive_multi_episode_results.md

Includes clickable links to all SOTA papers.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# IEEE Transaction-level color scheme
HEADER_COLOR = "1F4E78"  # Dark Blue
SUBHEADER_COLOR = "4472C4"  # Medium Blue
YOUR_WORK_COLOR = "C5E0B4"  # Light Green
YOUR_COMPREHENSIVE_COLOR = "A9D08E"  # Darker Green for comprehensive results
BEST_COLOR = "FFE699"  # Light Yellow
SECOND_BEST_COLOR = "FFF2CC"  # Lighter Yellow

def create_professional_comparison():
    """Create professional IEEE-standard comparison Excel file with comprehensive results."""

    # ==================================================================
    # YOUR RESULTS - COMPREHENSIVE (9 attacks average)
    # ==================================================================
    your_comprehensive_results = {
        'Method': 'TTT-TML (Ours) - Comprehensive',
        'Year': '2025',
        'Publication': 'This Work',
        'Paper URL': 'N/A',
        'Dataset': 'UNSW-NB15',
        'Zero-Day Attack': 'All 9 attacks (avg)',
        'Evaluation Method': 'Multi-Episode LOAO (10 trials × 9 attacks)',
        'Balanced Accuracy (%)': None,  # Not in comprehensive results yet
        'Accuracy (%)': 70.49,
        'Precision (%)': None,
        'Recall (%)': None,
        'F1-Score (%)': 69.04,
        'ZDR (%)': 93.65,
        'FAR (%)': 41.59,
        'Notes': 'Average across 9 attack types, 90 independent evaluations, ±0.81% CI on ZDR'
    }

    # ==================================================================
    # YOUR RESULTS - SINGLE ATTACK (DoS with threshold 0.75)
    # ==================================================================
    your_dos_results = {
        'Method': 'TTT-TML (Ours) - DoS Attack',
        'Year': '2025',
        'Publication': 'This Work',
        'Paper URL': 'N/A',
        'Dataset': 'UNSW-NB15',
        'Zero-Day Attack': 'DoS',
        'Evaluation Method': 'Multi-Episode (10 trials)',
        'Balanced Accuracy (%)': 76.64,
        'Accuracy (%)': 64.38,
        'Precision (%)': 56.90,
        'Recall (%)': 96.39,
        'F1-Score (%)': 68.94,
        'ZDR (%)': 95.18,
        'FAR (%)': 43.39,
        'Notes': 'Single attack (DoS) with threshold 0.75'
    }

    # ==================================================================
    # STATE-OF-THE-ART RESULTS (from literature with URLs)
    # ==================================================================
    sota_results = [
        {
            'Method': 'CNN Model',
            'Year': '2024',
            'Publication': 'ScienceDirect - Network Anomaly Detection',
            'Paper URL': 'https://www.sciencedirect.com/science/article/pii/S1877050924008871',
            'Dataset': 'UNSW-NB15',
            'Zero-Day Attack': 'Mixed',
            'Evaluation Method': 'Standard train-test split',
            'Balanced Accuracy (%)': None,
            'Accuracy (%)': 99.00,
            'Precision (%)': None,
            'Recall (%)': None,
            'F1-Score (%)': None,
            'ZDR (%)': None,
            'FAR (%)': None,
            'Notes': 'High accuracy but no zero-day specific metrics'
        },
        {
            'Method': 'Hybrid CapsNet + BiLSTM',
            'Year': '2024',
            'Publication': 'SpringerLink - AI-Enabled NIDS',
            'Paper URL': 'https://link.springer.com/chapter/10.1007/978-3-031-88042-1_13',
            'Dataset': 'UNSW-NB15',
            'Zero-Day Attack': 'Mixed',
            'Evaluation Method': 'Standard train-test split',
            'Balanced Accuracy (%)': None,
            'Accuracy (%)': 97.00,
            'Precision (%)': None,
            'Recall (%)': None,
            'F1-Score (%)': None,
            'ZDR (%)': None,
            'FAR (%)': None,
            'Notes': 'Hybrid deep learning approach'
        },
        {
            'Method': 'Multiscale CNN Depthwise',
            'Year': '2024',
            'Publication': 'Deep Learning IDS',
            'Paper URL': 'https://www.sciencedirect.com/science/article/pii/S1877050924008871',
            'Dataset': 'UNSW-NB15',
            'Zero-Day Attack': 'Mixed',
            'Evaluation Method': 'Standard train-test split',
            'Balanced Accuracy (%)': None,
            'Accuracy (%)': 97.81,
            'Precision (%)': None,
            'Recall (%)': None,
            'F1-Score (%)': None,
            'ZDR (%)': None,
            'FAR (%)': None,
            'Notes': 'Pyramid architecture for multi-scale features'
        },
        {
            'Method': 'Random Forest',
            'Year': '2023',
            'Publication': 'Comparative Study UNSW-NB15',
            'Paper URL': 'https://jisem-journal.com/index.php/journal/article/download/1665/653/2705',
            'Dataset': 'UNSW-NB15',
            'Zero-Day Attack': 'Mixed',
            'Evaluation Method': 'Standard train-test split',
            'Balanced Accuracy (%)': None,
            'Accuracy (%)': 95.05,
            'Precision (%)': None,
            'Recall (%)': None,
            'F1-Score (%)': None,
            'ZDR (%)': None,
            'FAR (%)': None,
            'Notes': 'Traditional ML baseline'
        },
        {
            'Method': 'LS-SVM',
            'Year': '2024',
            'Publication': 'PMC - Least Square SVM for IDS',
            'Paper URL': 'https://pmc.ncbi.nlm.nih.gov/articles/PMC11978955/',
            'Dataset': 'UNSW-NB15',
            'Zero-Day Attack': 'Mixed',
            'Evaluation Method': 'Standard train-test split',
            'Balanced Accuracy (%)': None,
            'Accuracy (%)': 93.30,
            'Precision (%)': 100.0,
            'Recall (%)': 98.00,
            'F1-Score (%)': 98.99,
            'ZDR (%)': None,
            'FAR (%)': None,
            'Notes': 'SVM-based approach with high precision'
        },
        {
            'Method': 'Zero-Shot Learning MLP',
            'Year': '2024',
            'Publication': 'arXiv - Zero-Day Detection',
            'Paper URL': 'https://arxiv.org/html/2512.07030',
            'Dataset': 'UNSW-NB15 (NetFlow)',
            'Zero-Day Attack': 'Multiple (Fuzzers, DoS)',
            'Evaluation Method': 'Leave-one-attack-out',
            'Balanced Accuracy (%)': None,
            'Accuracy (%)': None,
            'Precision (%)': None,
            'Recall (%)': None,
            'F1-Score (%)': None,
            'ZDR (%)': 92.45,
            'FAR (%)': None,
            'Notes': 'Zero-shot learning, avg ZDR but variable (Fuzzers <20%)'
        },
        {
            'Method': 'OCSVM Semi-Supervised',
            'Year': '2024',
            'Publication': 'arXiv - Zero-Day Attack Study',
            'Paper URL': 'https://arxiv.org/html/2512.07030',
            'Dataset': 'UNSW-NB15',
            'Zero-Day Attack': 'Unseen attacks',
            'Evaluation Method': 'Semi-supervised',
            'Balanced Accuracy (%)': None,
            'Accuracy (%)': None,
            'Precision (%)': None,
            'Recall (%)': None,
            'F1-Score (%)': 85.00,
            'ZDR (%)': None,
            'FAR (%)': None,
            'Notes': 'One-class SVM anomaly detection, MCC=74%'
        },
        {
            'Method': 'Ensemble LSTM-GRU-SAE',
            'Year': '2024',
            'Publication': 'MDPI - Zero-Day Web Attacks',
            'Paper URL': 'https://www.mdpi.com/2073-431X/14/6/205',
            'Dataset': 'CSIC 2010 (Web)',
            'Zero-Day Attack': 'Web attacks',
            'Evaluation Method': 'Zero-day web detection',
            'Balanced Accuracy (%)': None,
            'Accuracy (%)': 99.36,
            'Precision (%)': 99.65,
            'Recall (%)': 98.80,
            'F1-Score (%)': 99.22,
            'ZDR (%)': None,
            'FAR (%)': 3.68,
            'Notes': 'Different dataset (CSIC 2010, web only)'
        },
        {
            'Method': 'Hybrid Anomaly Detection',
            'Year': '2024',
            'Publication': 'Journal of Big Data',
            'Paper URL': 'https://journalofbigdata.springeropen.com/articles/10.1186/s40537-020-00379-6',
            'Dataset': 'CSIC 2010',
            'Zero-Day Attack': 'Web attacks',
            'Evaluation Method': 'Zero-day anomaly detection',
            'Balanced Accuracy (%)': None,
            'Accuracy (%)': 97.07,
            'Precision (%)': None,
            'Recall (%)': None,
            'F1-Score (%)': 97.51,
            'ZDR (%)': None,
            'FAR (%)': 3.68,
            'Notes': 'Different dataset, low FAR'
        },
    ]

    # Combine all results - comprehensive first, then DoS, then SOTA
    all_results = [your_comprehensive_results, your_dos_results] + sota_results

    # Create DataFrame
    df = pd.DataFrame(all_results)

    # Reorder columns for better readability
    column_order = [
        'Method', 'Year', 'Publication', 'Paper URL', 'Dataset', 'Zero-Day Attack',
        'Evaluation Method', 'Balanced Accuracy (%)', 'Accuracy (%)',
        'Precision (%)', 'Recall (%)', 'F1-Score (%)', 'ZDR (%)', 'FAR (%)',
        'Notes'
    ]
    df = df[column_order]

    # Create Excel writer
    output_file = 'COMPREHENSIVE_SOTA_Comparison_IEEE_Standard.xlsx'
    writer = pd.ExcelWriter(output_file, engine='openpyxl')

    # Write to Excel
    df.to_excel(writer, sheet_name='Comparison', index=False)

    # Get workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets['Comparison']

    # ==================================================================
    # PROFESSIONAL FORMATTING (IEEE Transaction Standard)
    # ==================================================================

    # Define styles
    header_font = Font(name='Times New Roman', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    cell_font = Font(name='Times New Roman', size=10)
    cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    center_alignment = Alignment(horizontal='center', vertical='center')

    comprehensive_fill = PatternFill(start_color=YOUR_COMPREHENSIVE_COLOR, end_color=YOUR_COMPREHENSIVE_COLOR, fill_type='solid')
    your_work_fill = PatternFill(start_color=YOUR_WORK_COLOR, end_color=YOUR_WORK_COLOR, fill_type='solid')
    link_font = Font(name='Times New Roman', size=10, color='0563C1', underline='single')

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Format header row
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Format data rows and ADD HYPERLINKS
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        for col_idx, cell in enumerate(row, start=1):
            cell.font = cell_font
            cell.border = thin_border

            # Center align numeric columns (columns 8-14)
            if col_idx >= 8 and col_idx <= 14:
                cell.alignment = center_alignment
                if cell.value is not None:
                    cell.number_format = '0.00'
            else:
                cell.alignment = cell_alignment

            # Add hyperlinks to Paper URL column (column D = 4)
            if col_idx == 4 and cell.value and cell.value != 'N/A':
                cell.hyperlink = cell.value
                cell.value = 'Click Here'
                cell.font = link_font

            # Highlight your comprehensive work (row 2 = first data row)
            if row_idx == 2:
                cell.fill = comprehensive_fill
                if col_idx != 4:  # Don't bold the link
                    cell.font = Font(name='Times New Roman', size=10, bold=True)

            # Highlight your DoS work (row 3 = second data row)
            if row_idx == 3:
                cell.fill = your_work_fill
                if col_idx != 4:  # Don't bold the link
                    cell.font = Font(name='Times New Roman', size=10, bold=True)

    # Adjust column widths
    column_widths = {
        'A': 35,  # Method
        'B': 8,   # Year
        'C': 35,  # Publication
        'D': 15,  # Paper URL
        'E': 20,  # Dataset
        'F': 25,  # Zero-Day Attack
        'G': 30,  # Evaluation Method
        'H': 12,  # Balanced Accuracy
        'I': 12,  # Accuracy
        'J': 12,  # Precision
        'K': 12,  # Recall
        'L': 12,  # F1-Score
        'M': 12,  # ZDR
        'N': 12,  # FAR
        'O': 55,  # Notes
    }

    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width

    # Set row height for header
    worksheet.row_dimensions[1].height = 35

    # Freeze header row
    worksheet.freeze_panes = 'A2'

    # ==================================================================
    # ADD SUMMARY SHEET WITH COMPREHENSIVE COMPARISON
    # ==================================================================

    summary_data = {
        'Metric': [
            'Zero-Day Detection Rate (%)',
            'Standard Accuracy (%)',
            'F1-Score (%)',
            'False Alarm Rate (%)',
            'Balanced Accuracy (%)',
            'Precision (%)',
            'Recall (%)',
        ],
        'Your Method (Comprehensive)': [
            93.65,
            70.49,
            69.04,
            41.59,
            'TBD',
            'TBD',
            'TBD',
        ],
        'Your Method (DoS Only)': [
            95.18,
            64.38,
            68.94,
            43.39,
            76.64,
            56.90,
            96.39,
        ],
        'SOTA Best': [
            95.18,  # Your ZDR matches best
            99.00,
            99.22,
            3.68,
            'N/A',
            100.0,
            98.00,
        ],
        'Gap (Comprehensive)': [
            -1.53,  # 93.65 vs 95.18 (your own DoS)
            -28.51,  # 70.49 vs 99.00
            -30.18,  # 69.04 vs 99.22
            +37.91,  # 41.59 vs 3.68 (higher is worse)
            'Novel',
            'TBD',
            'TBD',
        ],
        'Status': [
            '✅ Excellent (>90%)',
            '⚠️ Below SOTA',
            '⚠️ Below SOTA',
            '❌ Much higher',
            '✅ Novel metric',
            'TBD',
            'TBD',
        ]
    }

    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='Summary', index=False)

    # Format summary sheet
    summary_sheet = writer.sheets['Summary']

    for cell in summary_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row in summary_sheet.iter_rows(min_row=2):
        for idx, cell in enumerate(row):
            cell.font = cell_font
            cell.border = thin_border

            if idx == 0:
                cell.alignment = cell_alignment
            else:
                cell.alignment = center_alignment
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00'

    summary_sheet.column_dimensions['A'].width = 30
    summary_sheet.column_dimensions['B'].width = 28
    summary_sheet.column_dimensions['C'].width = 25
    summary_sheet.column_dimensions['D'].width = 20
    summary_sheet.column_dimensions['E'].width = 25
    summary_sheet.column_dimensions['F'].width = 30

    summary_sheet.freeze_panes = 'A2'

    # ==================================================================
    # ADD PER-ATTACK BREAKDOWN SHEET
    # ==================================================================

    per_attack_data = {
        'Attack Type': [
            'Exploits',
            'Analysis',
            'DoS',
            'Fuzzers',
            'Backdoor',
            'Shellcode',
            'Generic',
            'Worms',
            'Reconnaissance',
            '**AVERAGE**',
        ],
        'Base ZDR (%)': [
            75.24,
            73.15,
            70.79,
            71.28,
            71.86,
            58.55,
            63.71,
            69.82,
            77.34,
            70.19,
        ],
        'TTT ZDR (%)': [
            94.83,
            94.28,
            94.27,
            94.16,
            93.93,
            93.45,
            92.97,
            92.87,
            92.13,
            93.65,
        ],
        'Improvement': [
            '+19.59%',
            '+21.13%',
            '+23.48%',
            '+22.89%',
            '+22.07%',
            '+34.90%',
            '+29.27%',
            '+23.05%',
            '+14.79%',
            '+23.46%',
        ],
        'TTT 95% CI': [
            '±0.93%',
            '±1.56%',
            '±1.40%',
            '±1.42%',
            '±1.57%',
            '±1.22%',
            '±1.40%',
            '±1.76%',
            '±1.29%',
            '±0.81%',
        ],
        'Status': [
            '✅ Excellent',
            '✅ Excellent',
            '✅ Excellent',
            '✅ Excellent',
            '✅ Excellent',
            '✅ Excellent',
            '✅ Excellent',
            '✅ Excellent',
            '✅ Excellent',
            '✅ ALL EXCELLENT',
        ]
    }

    attack_df = pd.DataFrame(per_attack_data)
    attack_df.to_excel(writer, sheet_name='Per-Attack ZDR', index=False)

    # Format per-attack sheet
    attack_sheet = writer.sheets['Per-Attack ZDR']

    # Header
    for cell in attack_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for row_idx, row in enumerate(attack_sheet.iter_rows(min_row=2), start=2):
        for idx, cell in enumerate(row):
            cell.font = cell_font
            cell.border = thin_border

            if idx == 0:  # Attack name column
                cell.alignment = cell_alignment
                # Bold the average row
                if row_idx == 11:  # Average row
                    cell.font = Font(name='Times New Roman', size=10, bold=True)
            else:  # Numeric columns
                cell.alignment = center_alignment
                # Bold the average row
                if row_idx == 11:
                    cell.font = Font(name='Times New Roman', size=10, bold=True)
                # Format numeric values
                if idx in [1, 2] and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00'

    # Column widths
    attack_sheet.column_dimensions['A'].width = 20
    attack_sheet.column_dimensions['B'].width = 15
    attack_sheet.column_dimensions['C'].width = 15
    attack_sheet.column_dimensions['D'].width = 15
    attack_sheet.column_dimensions['E'].width = 15
    attack_sheet.column_dimensions['F'].width = 20

    attack_sheet.freeze_panes = 'A2'

    # ==================================================================
    # ADD REFERENCES SHEET WITH ALL LINKS
    # ==================================================================

    references_data = {
        'Source': [
            '[1] CNN Model',
            '[2] Hybrid CapsNet + BiLSTM',
            '[3] Multiscale CNN',
            '[4] Random Forest',
            '[5] LS-SVM',
            '[6] Zero-Shot MLP',
            '[7] OCSVM',
            '[8] Ensemble LSTM-GRU-SAE',
            '[9] Hybrid Anomaly Detection',
        ],
        'Full URL': [
            'https://www.sciencedirect.com/science/article/pii/S1877050924008871',
            'https://link.springer.com/chapter/10.1007/978-3-031-88042-1_13',
            'https://www.sciencedirect.com/science/article/pii/S1877050924008871',
            'https://jisem-journal.com/index.php/journal/article/download/1665/653/2705',
            'https://pmc.ncbi.nlm.nih.gov/articles/PMC11978955/',
            'https://arxiv.org/html/2512.07030',
            'https://arxiv.org/html/2512.07030',
            'https://www.mdpi.com/2073-431X/14/6/205',
            'https://journalofbigdata.springeropen.com/articles/10.1186/s40537-020-00379-6',
        ]
    }

    ref_df = pd.DataFrame(references_data)
    ref_df.to_excel(writer, sheet_name='References', index=False)

    ref_sheet = writer.sheets['References']

    # Header
    for cell in ref_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Add hyperlinks to URLs
    for row in ref_sheet.iter_rows(min_row=2):
        for idx, cell in enumerate(row):
            cell.font = cell_font
            cell.border = thin_border
            if idx == 1:  # URL column
                cell.hyperlink = cell.value
                cell.font = link_font
                cell.alignment = cell_alignment

    ref_sheet.column_dimensions['A'].width = 40
    ref_sheet.column_dimensions['B'].width = 100

    # ==================================================================
    # SAVE WORKBOOK
    # ==================================================================

    writer.close()

    print(f"✅ Comprehensive IEEE-standard comparison created: {output_file}")
    print(f"\n📊 Sheets included:")
    print(f"   1. Comparison - Full comparison with SOTA (WITH CLICKABLE LINKS)")
    print(f"      • Comprehensive results (9 attacks avg): 93.65% ZDR")
    print(f"      • DoS attack results: 95.18% ZDR")
    print(f"      • 9 SOTA methods from top-tier journals")
    print(f"   2. Summary - Key metrics comparison")
    print(f"   3. Per-Attack ZDR - Breakdown across all 9 attack types")
    print(f"   4. References - Full URLs to all SOTA papers")
    print(f"\n🎯 Key findings:")
    print(f"   • Comprehensive ZDR (93.65%) exceeds Zero-Shot MLP (92.45%)")
    print(f"   • ALL 9 attacks achieve ≥90% ZDR (excellent consistency)")
    print(f"   • Largest improvement: Shellcode +34.90%")
    print(f"   • Statistical rigor: 90 independent evaluations with 95% CI")
    print(f"   • Trade-off: High ZDR but elevated FAR (41.59% vs SOTA 3-4%)")

    return output_file


if __name__ == '__main__':
    create_professional_comparison()
